"""
Tiny shared .xlsx helper for every admin import/export screen (Expense
Categories, Users, and anything added later) - so each one doesn't hand-roll
its own openpyxl boilerplate. Deliberately dumb: no styling opinions beyond a
bold header row, no schema validation - callers own their own column list
and row parsing.
"""

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_workbook(headers: list[str], rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def read_workbook(file_stream) -> tuple[list[str], list[dict]]:
    """Returns (original-case headers, rows) - each row is a dict keyed by
    the LOWERCASED, stripped header, so callers can look up row.get("name")
    regardless of how the header cell was capitalized in a hand-edited file.
    Fully blank rows (every cell empty) are dropped."""
    # openpyxl/zipfile call .seekable() on whatever's handed in, which a
    # Werkzeug upload's underlying SpooledTemporaryFile doesn't implement -
    # wrapping in BytesIO (which always supports it) sidesteps that instead
    # of relying on every caller's upload stream happening to support it.
    wb = load_workbook(io.BytesIO(file_stream.read()), data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    keys = [h.lower() for h in headers]

    rows = []
    for raw in rows_iter:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row = {}
        for key, value in zip(keys, raw):
            if key:
                row[key] = value
        rows.append(row)

    return headers, rows
